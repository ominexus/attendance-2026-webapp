#!/usr/bin/env python3
"""
2026 고등부 출석부 엑셀 → Supabase 마이그레이션 스크립트
Usage: python3 migrate.py [--excel-path <path>] [--dry-run]
"""

import argparse
import sys
import os
import json
from datetime import datetime, date
import openpyxl
import requests

# ─── 설정 ───────────────────────────────────────────────────────────────────
SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://ovtgwbhbwtfwzgaihlmb.supabase.co")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")  # 반드시 환경변수로 주입 (service_role 키 권장)
EXCEL_PATH = os.environ.get("EXCEL_PATH", "2026 고등부 출석부.xlsx")

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}


def supabase_insert(table: str, rows: list[dict], dry_run: bool = False) -> list[dict]:
    """Supabase REST API를 통해 데이터를 삽입하고 삽입된 행을 반환합니다."""
    if not rows:
        return []
    if dry_run:
        print(f"  [DRY-RUN] {table}: {len(rows)}건 삽입 예정")
        return rows
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    resp = requests.post(url, headers=HEADERS, json=rows, timeout=30)
    if resp.status_code not in (200, 201):
        print(f"  [ERROR] {table} 삽입 실패: {resp.status_code} {resp.text[:300]}")
        sys.exit(1)
    return resp.json()


def supabase_select(table: str, select: str = "*") -> list[dict]:
    """Supabase REST API를 통해 데이터를 조회합니다."""
    url = f"{SUPABASE_URL}/rest/v1/{table}?select={select}"
    resp = requests.get(url, headers=HEADERS, timeout=30)
    if resp.status_code != 200:
        print(f"  [ERROR] {table} 조회 실패: {resp.status_code} {resp.text[:300]}")
        sys.exit(1)
    return resp.json()


def parse_date(val) -> str | None:
    """다양한 형식의 날짜 값을 ISO 8601 문자열로 변환합니다."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str):
        for fmt in ("%Y.%m.%d", "%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(val.strip(), fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return None


def load_workbook(path: str):
    """엑셀 파일을 로드합니다."""
    if not os.path.exists(path):
        print(f"[ERROR] 엑셀 파일을 찾을 수 없습니다: {path}")
        sys.exit(1)
    return openpyxl.load_workbook(path, data_only=True)


def parse_teachers(wb) -> list[dict]:
    """교사 명단 시트를 파싱합니다."""
    ws = wb["교사 명단"]
    teachers = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        role_cell = row[1].value  # B열: 직분
        name_cell = row[2].value  # C열: 이름
        if role_cell and name_cell:
            teachers.append({
                "role": str(role_cell).strip(),
                "name": str(name_cell).strip(),
            })
    return teachers


def parse_students(wb) -> list[dict]:
    """학생 명단 시트를 파싱합니다."""
    ws = wb["학생 명단"]
    students = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        grade = row[1].value  # B열: 학년
        class_num = row[2].value  # C열: 반
        name = row[3].value  # D열: 이름
        if not (grade and class_num and name):
            continue
        gender = row[4].value if row[4].value else None  # E열: 성별
        phone = str(row[5].value).strip() if row[5].value else None  # F열: 전화번호
        birth_date = parse_date(row[6].value)  # G열: 생일
        school = str(row[7].value).strip() if row[7].value else None  # H열: 학교
        guide = str(row[8].value).strip() if row[8].value else None  # I열: 인도자
        students.append({
            "grade": str(grade).strip(),
            "class_num": str(class_num).strip(),
            "name": str(name).strip(),
            "gender": str(gender).strip() if gender else None,
            "phone": phone,
            "birth_date": birth_date,
            "school": school,
            "guide": guide,
        })
    return students


def parse_attendance(wb, student_id_map: dict) -> list[dict]:
    """
    전체학생 시트를 파싱하여 출석 기록을 생성합니다.
    student_id_map: {(grade, class_num, name): uuid} 형태
    """
    ws = wb["전체학생"]
    attendance_records = []

    # 헤더 행(2행)에서 날짜 컬럼 위치 파악 (E열부터 BD열까지)
    header_row = list(ws.iter_rows(min_row=2, max_row=2))[0]
    date_cols = {}  # col_idx → date string
    for cell in header_row:
        if isinstance(cell.value, (datetime, date)):
            date_cols[cell.column] = cell.value.strftime("%Y-%m-%d")

    # 3행부터 학생 데이터 파싱
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        grade = row[0].value  # A열: 학년
        class_num = row[1].value  # B열: 반
        name = row[2].value  # C열: 이름
        if not (grade and class_num and name):
            continue
        key = (str(grade).strip(), str(class_num).strip(), str(name).strip())
        student_id = student_id_map.get(key)
        if not student_id:
            print(f"  [WARN] 학생 ID 미발견: {key}")
            continue
        for cell in row:
            if cell.column not in date_cols:
                continue
            att_date = date_cols[cell.column]
            status = bool(cell.value) if cell.value is not None else False
            attendance_records.append({
                "student_id": student_id,
                "attendance_date": att_date,
                "status": status,
            })
    return attendance_records


def main():
    parser = argparse.ArgumentParser(description="2026 고등부 출석부 Supabase 마이그레이션")
    parser.add_argument("--excel-path", default=EXCEL_PATH, help="엑셀 파일 경로")
    parser.add_argument("--dry-run", action="store_true", help="실제 삽입 없이 파싱 결과만 출력")
    args = parser.parse_args()

    if not args.dry_run and not SUPABASE_KEY:
        print("[ERROR] SUPABASE_KEY 환경변수가 설정되지 않았습니다.")
        print("  export SUPABASE_KEY=<service_role_key>")
        sys.exit(1)

    print(f"[1/4] 엑셀 파일 로드: {args.excel_path}")
    wb = load_workbook(args.excel_path)

    # ── 교사 명단 마이그레이션 ─────────────────────────────────────────────
    print("[2/4] 교사 명단 파싱 및 삽입...")
    teachers = parse_teachers(wb)
    print(f"  파싱 완료: {len(teachers)}명")
    inserted_teachers = supabase_insert("teachers", teachers, dry_run=args.dry_run)
    print(f"  삽입 완료: {len(inserted_teachers)}건")

    # ── 학생 명단 마이그레이션 ─────────────────────────────────────────────
    print("[3/4] 학생 명단 파싱 및 삽입...")
    students = parse_students(wb)
    print(f"  파싱 완료: {len(students)}명")
    inserted_students = supabase_insert("students", students, dry_run=args.dry_run)
    print(f"  삽입 완료: {len(inserted_students)}건")

    # 삽입된 학생 UUID 매핑 (dry-run 시 임시 UUID 사용)
    if args.dry_run:
        import uuid
        student_id_map = {
            (s["grade"], s["class_num"], s["name"]): str(uuid.uuid4())
            for s in students
        }
    else:
        all_students = supabase_select("students", "id,grade,class_num,name")
        student_id_map = {
            (s["grade"], s["class_num"], s["name"]): s["id"]
            for s in all_students
        }

    # ── 출석 기록 마이그레이션 ─────────────────────────────────────────────
    print("[4/4] 출석 기록 파싱 및 삽입...")
    attendance_records = parse_attendance(wb, student_id_map)
    # status=True인 기록만 삽입 (False는 기본값이므로 생략 가능, 필요 시 모두 삽입)
    true_records = [r for r in attendance_records if r["status"]]
    print(f"  파싱 완료: 전체 {len(attendance_records)}건 / 출석(True) {len(true_records)}건")

    # 배치 단위로 삽입 (Supabase REST API 요청 크기 제한 대비)
    BATCH_SIZE = 500
    total_inserted = 0
    for i in range(0, len(true_records), BATCH_SIZE):
        batch = true_records[i:i + BATCH_SIZE]
        inserted = supabase_insert("attendance", batch, dry_run=args.dry_run)
        total_inserted += len(inserted)
        print(f"  배치 {i // BATCH_SIZE + 1}: {len(inserted)}건 삽입")
    print(f"  삽입 완료: {total_inserted}건")

    print("\n[완료] 마이그레이션이 성공적으로 완료되었습니다.")


if __name__ == "__main__":
    main()
