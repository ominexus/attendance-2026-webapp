#!/usr/bin/env python3
"""
MCP execute_sql을 통한 Supabase 직접 마이그레이션 스크립트
service_role 키 없이 MCP 도구를 사용하여 데이터를 삽입합니다.
"""

import subprocess
import json
import sys
import os
from datetime import datetime, date
import openpyxl

EXCEL_PATH = "/home/ubuntu/attendance/2026 고등부 출석부.xlsx"
PROJECT_ID = "ovtgwbhbwtfwzgaihlmb"


def run_mcp_sql(sql: str) -> dict:
    """MCP execute_sql 도구를 호출합니다."""
    payload = json.dumps({"project_id": PROJECT_ID, "query": sql})
    result = subprocess.run(
        ["manus-mcp-cli", "tool", "call", "execute_sql", "--server", "supabase", "--input", payload],
        capture_output=True, text=True, timeout=60
    )
    if result.returncode != 0:
        print(f"[ERROR] MCP 호출 실패: {result.stderr[:200]}")
        return {}
    # 결과에서 JSON 파싱
    output = result.stdout
    try:
        # "Tool execution result:" 이후의 JSON 추출
        idx = output.find("Tool execution result:\n")
        if idx >= 0:
            json_str = output[idx + len("Tool execution result:\n"):].strip()
            return json.loads(json_str)
    except Exception as e:
        print(f"[WARN] 결과 파싱 실패: {e}")
    return {}


def escape_sql_string(val) -> str:
    """SQL 문자열 이스케이프 처리."""
    if val is None:
        return "NULL"
    return "'" + str(val).replace("'", "''") + "'"


def parse_date(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str):
        for fmt in ("%Y.%m.%d", "%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(val.strip(), fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return None


def main():
    print(f"[1/4] 엑셀 파일 로드: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # ── 교사 명단 삽입 ─────────────────────────────────────────────────────
    print("[2/4] 교사 명단 삽입...")
    ws = wb["교사 명단"]
    teacher_values = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        role = row[1].value
        name = row[2].value
        if role and name:
            teacher_values.append(f"({escape_sql_string(str(role).strip())}, {escape_sql_string(str(name).strip())})")

    if teacher_values:
        sql = f"INSERT INTO public.teachers (role, name) VALUES {', '.join(teacher_values)} ON CONFLICT DO NOTHING;"
        result = run_mcp_sql(sql)
        print(f"  교사 {len(teacher_values)}명 삽입 완료")

    # ── 학생 명단 삽입 ─────────────────────────────────────────────────────
    print("[3/4] 학생 명단 삽입...")
    ws = wb["학생 명단"]
    student_values = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        grade = row[1].value
        class_num = row[2].value
        name = row[3].value
        if not (grade and class_num and name):
            continue
        gender = row[4].value
        phone = row[5].value
        birth_date = parse_date(row[6].value)
        school = row[7].value
        guide = row[8].value
        student_values.append(
            f"({escape_sql_string(str(grade).strip())}, "
            f"{escape_sql_string(str(class_num).strip())}, "
            f"{escape_sql_string(str(name).strip())}, "
            f"{escape_sql_string(str(gender).strip()) if gender else 'NULL'}, "
            f"{escape_sql_string(str(phone).strip()) if phone else 'NULL'}, "
            f"{escape_sql_string(birth_date) if birth_date else 'NULL'}, "
            f"{escape_sql_string(str(school).strip()) if school else 'NULL'}, "
            f"{escape_sql_string(str(guide).strip()) if guide else 'NULL'})"
        )

    if student_values:
        sql = (
            "INSERT INTO public.students (grade, class_num, name, gender, phone, birth_date, school, guide) "
            f"VALUES {', '.join(student_values)} ON CONFLICT DO NOTHING;"
        )
        result = run_mcp_sql(sql)
        print(f"  학생 {len(student_values)}명 삽입 완료")

    # ── 학생 UUID 조회 ─────────────────────────────────────────────────────
    print("  학생 UUID 조회 중...")
    result = run_mcp_sql("SELECT id, grade, class_num, name FROM public.students;")
    student_id_map = {}
    if result and "result" in result:
        # untrusted-data 태그 내 JSON 추출
        raw = result["result"]
        start = raw.find("[{")
        end = raw.rfind("}]") + 2
        if start >= 0 and end > start:
            rows = json.loads(raw[start:end])
            for r in rows:
                student_id_map[(r["grade"], r["class_num"], r["name"])] = r["id"]
    print(f"  UUID 매핑 완료: {len(student_id_map)}명")

    # ── 출석 기록 삽입 ─────────────────────────────────────────────────────
    print("[4/4] 출석 기록 삽입...")
    ws = wb["전체학생"]
    header_row = list(ws.iter_rows(min_row=2, max_row=2))[0]
    date_cols = {}
    for cell in header_row:
        if isinstance(cell.value, (datetime, date)):
            date_cols[cell.column] = cell.value.strftime("%Y-%m-%d")

    att_values = []
    warn_count = 0
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        grade = row[0].value
        class_num = row[1].value
        name = row[2].value
        if not (grade and class_num and name):
            continue
        key = (str(grade).strip(), str(class_num).strip(), str(name).strip())
        student_id = student_id_map.get(key)
        if not student_id:
            warn_count += 1
            continue
        for cell in row:
            if cell.column not in date_cols:
                continue
            status = bool(cell.value) if cell.value is not None else False
            if not status:
                continue  # False(결석) 기록은 생략 (기본값이므로)
            att_values.append(
                f"({escape_sql_string(student_id)}, "
                f"{escape_sql_string(date_cols[cell.column])}, "
                f"TRUE)"
            )

    if warn_count:
        print(f"  [WARN] UUID 미발견 학생: {warn_count}명")

    # 배치 삽입 (500건씩)
    BATCH_SIZE = 500
    total = 0
    for i in range(0, len(att_values), BATCH_SIZE):
        batch = att_values[i:i + BATCH_SIZE]
        sql = (
            "INSERT INTO public.attendance (student_id, attendance_date, status) "
            f"VALUES {', '.join(batch)} ON CONFLICT (student_id, attendance_date) DO UPDATE SET status = EXCLUDED.status;"
        )
        run_mcp_sql(sql)
        total += len(batch)
        print(f"  배치 {i // BATCH_SIZE + 1}: {len(batch)}건 삽입 ({total}/{len(att_values)})")

    print(f"\n[완료] 마이그레이션 완료 - 교사 {len(teacher_values)}명, 학생 {len(student_values)}명, 출석 {total}건")


if __name__ == "__main__":
    main()
